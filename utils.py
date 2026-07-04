# utils.py — Excel export with embedded photos
import os
import pandas as pd
from datetime import datetime


# ────────────────────────────────────────────────────────── #
#  Excel master file with openpyxl (photos embedded in rows) #
# ────────────────────────────────────────────────────────── #

EXCEL_PATH = "Attendance_Master.xlsx"
PHOTO_COL   = "G"
PHOTO_W     = 60   # px
PHOTO_H     = 60   # px
ROW_H_PTS   = 50   # Excel row height (points ≈ 0.75 * px)

HEADERS = ["S.No", "Student ID", "Name", "Date", "Time", "Status", "Photo"]


def _ensure_workbook():
    """Return (wb, ws) creating the file with headers if it doesn't exist."""
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    if os.path.exists(EXCEL_PATH):
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"

        # Header row styling
        header_fill = PatternFill("solid", fgColor="1F6FEB")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for col, h in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Column widths
        widths = [8, 12, 20, 14, 12, 10, 12]
        from openpyxl.utils import get_column_letter
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    return wb, ws


def update_excel(stu_id: int, name: str, date: str, time_str: str,
                 status: str = "Present", photo_path: str = None):
    """Append one attendance row, embedding the face photo if available."""
    try:
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Alignment

        wb, ws = _ensure_workbook()

        row_num = ws.max_row + 1
        sno = row_num - 1  # row 1 is headers

        ws.cell(row_num, 1, sno)
        ws.cell(row_num, 2, stu_id)
        ws.cell(row_num, 3, name)
        ws.cell(row_num, 4, date)
        ws.cell(row_num, 5, time_str)
        ws.cell(row_num, 6, status)

        # Center-align all data cells
        for col in range(1, 7):
            ws.cell(row_num, col).alignment = Alignment(
                horizontal="center", vertical="center"
            )

        # Embed photo
        if photo_path and os.path.exists(photo_path):
            ws.row_dimensions[row_num].height = ROW_H_PTS
            img = XLImage(photo_path)
            img.width  = PHOTO_W
            img.height = PHOTO_H
            ws.add_image(img, f"{PHOTO_COL}{row_num}")
        else:
            ws.cell(row_num, 7, "—")

        wb.save(EXCEL_PATH)

    except Exception as e:
        print(f"[Excel] Could not update master Excel: {e}")


def update_csv(stu_id: int, name: str, date: str, time_str: str,
               status: str = "Present"):
    """Append to daily CSV."""
    csv_path = f"attendance/Attendance_{date}.csv"
    os.makedirs("attendance", exist_ok=True)

    new_row = {
        "S.No": None,       # filled below
        "Student ID": stu_id,
        "Name": name,
        "Date": date,
        "Time": time_str,
        "Status": status
    }

    if os.path.exists(csv_path):
        df = pd.read_csv(csv_path)
    else:
        df = pd.DataFrame(columns=list(new_row.keys()))

    new_row["S.No"] = len(df) + 1
    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    df.to_csv(csv_path, index=False)


def export_daily_excel(date: str, dst_path: str):
    """Export a single day's attendance to a standalone Excel file."""
    from db import get_today_attendance
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter

    rows = get_today_attendance(date)
    wb = Workbook()
    ws = wb.active
    ws.title = f"Attendance {date}"

    headers = ["S.No", "Student ID", "Name", "Date", "Time", "Status", "Photo"]
    hfill = PatternFill("solid", fgColor="1F6FEB")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")

    widths = [8, 12, 20, 14, 12, 10, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for sno, r in enumerate(rows, 1):
        rn = sno + 1
        ws.cell(rn, 1, sno)
        ws.cell(rn, 2, r["student_id"])
        ws.cell(rn, 3, r["name"])
        ws.cell(rn, 4, date)
        ws.cell(rn, 5, r["time"])
        ws.cell(rn, 6, r.get("status", "Present"))
        for col in range(1, 7):
            ws.cell(rn, col).alignment = Alignment(
                horizontal="center", vertical="center"
            )
        photo = r.get("photo_path")
        if photo and os.path.exists(photo):
            ws.row_dimensions[rn].height = ROW_H_PTS
            img = XLImage(photo)
            img.width = PHOTO_W; img.height = PHOTO_H
            ws.add_image(img, f"G{rn}")
        else:
            ws.cell(rn, 7, "—")

    wb.save(dst_path)
